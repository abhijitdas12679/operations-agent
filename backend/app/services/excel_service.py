import re
from io import BytesIO
from openpyxl import load_workbook

EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")

REQUIRED_COLUMNS = {"name", "email", "designation"}


def parse_bulk_email_excel(file_bytes: bytes):
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Excel file is empty.")

    headers = [str(h or "").strip().lower() for h in rows[0]]
    missing = REQUIRED_COLUMNS - set(headers)

    if missing:
        raise ValueError(
            f"Missing required columns: {', '.join(missing)}. Required columns: name, email, designation"
        )

    recipients = []

    for index, row in enumerate(rows[1:], start=2):
        data = {}

        for i, value in enumerate(row):
            if i < len(headers):
                data[headers[i]] = str(value or "").strip()

        name = data.get("name", "")
        email = data.get("email", "")
        designation = data.get("designation", "")

        if not name and not email:
            continue

        if not name:
            raise ValueError(f"Name is missing at row {index}")

        if not email:
            raise ValueError(f"Email is missing at row {index}")

        if not EMAIL_RE.match(email):
            raise ValueError(f"Invalid email at row {index}: {email}")

        if not designation:
            raise ValueError(f"Designation is missing at row {index}")

        recipients.append({
            "name": name,
            "email": email,
            "designation": designation,
        })

    if not recipients:
        raise ValueError("No valid recipients found in Excel file.")

    return recipients