import json
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.auth import get_current_user
from app.crews.crew_manager import run_meeting_crew
from app.tools.pdf_tool import export_to_pdf
from app.services.smtp_service import smtp_service

router = APIRouter(prefix="/meeting", tags=["Meeting"])


def clean_text(text: str = "") -> str:
    return (
        text.replace("**", "")
        .replace("###", "##")
        .replace("---", "")
        .strip()
    )


def format_mom_output(meeting_title: str, attendees: str, generated: str) -> str:
    generated = clean_text(generated)

    return f"""# Minutes of Meeting

## Meeting Title
{meeting_title}

## Attendees
{attendees}

## Meeting Summary
{generated}
"""


def build_mom_email(
    recipient_name: str,
    position: str,
    meeting_title: str,
    sender_name: str | None = None,
    sender_designation: str | None = None,
) -> str:
    name = recipient_name.strip() if recipient_name else "Sir/Madam"
    position_text = f" ({position.strip()})" if position else ""

    return f"""Dear {name}{position_text},

Please find attached the Minutes of Meeting for "{meeting_title}".

The attached PDF includes the meeting summary, discussion points, decisions, and action items.

Kindly review the attached document.

Best regards,
{sender_name or "Team"}
{sender_designation or ""}
"""


@router.post("/generate-mom", response_model=schemas.MeetingMOMResponse, status_code=201)
def generate_mom(
    req: schemas.MeetingMOMRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    try:
        generated = run_meeting_crew(req.meeting_title, req.attendees, req.raw_notes)
        generated = format_mom_output(req.meeting_title, req.attendees, generated)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI generation failed: {str(e)}")

    record = models.MeetingHistory(
        user_id=current_user.id,
        meeting_title=req.meeting_title,
        attendees=req.attendees,
        raw_notes=req.raw_notes,
        generated_mom=generated,
    )

    db.add(record)
    db.commit()
    db.refresh(record)

    return record


@router.get("/history")
def get_meeting_history(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    return (
        db.query(models.MeetingHistory)
        .filter(models.MeetingHistory.user_id == current_user.id)
        .order_by(models.MeetingHistory.created_at.desc())
        .limit(50)
        .all()
    )


@router.get("/history/{meeting_id}")
def get_meeting_detail(
    meeting_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    record = (
        db.query(models.MeetingHistory)
        .filter(
            models.MeetingHistory.id == meeting_id,
            models.MeetingHistory.user_id == current_user.id,
        )
        .first()
    )

    if not record:
        raise HTTPException(status_code=404, detail="Meeting not found")

    return record


@router.post("/send")
def send_mom_manual(
    meeting_id: int = Form(...),
    recipients: str = Form(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    meeting = (
        db.query(models.MeetingHistory)
        .filter(
            models.MeetingHistory.id == meeting_id,
            models.MeetingHistory.user_id == current_user.id,
        )
        .first()
    )

    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting MOM not found")

    if not current_user.smtp_setting:
        raise HTTPException(status_code=400, detail="Please connect SMTP first.")

    try:
        recipient_list = json.loads(recipients)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid recipients data.")

    valid_recipients = []

    for item in recipient_list:
        email = str(item.get("email", "")).strip()
        if email:
            valid_recipients.append({
                "name": str(item.get("name", "")).strip() or "Sir/Madam",
                "position": str(item.get("position", "")).strip(),
                "email": email,
            })

    if not valid_recipients:
        raise HTTPException(status_code=400, detail="Please add at least one email.")

    pdf_path = export_to_pdf(
        meeting.generated_mom,
        "meeting",
        f"MOM - {meeting.meeting_title}",
    )

    subject = f"Minutes of Meeting - {meeting.meeting_title}"

    sent = 0
    failed = []

    for recipient in valid_recipients:
        try:
            smtp_service.send_email(
                smtp_setting=current_user.smtp_setting,
                to_email=recipient["email"],
                subject=subject,
                body=build_mom_email(
                    recipient_name=recipient["name"],
                    position=recipient["position"],
                    meeting_title=meeting.meeting_title,
                    sender_name=current_user.full_name,
                    sender_designation=current_user.designation,
                ),
                attachments=[pdf_path],
            )
            sent += 1
        except Exception as e:
            failed.append({"email": recipient["email"], "error": str(e)})

    return {"message": "MOM sending completed.", "sent": sent, "failed": failed}


@router.post("/send-excel")
def send_mom_excel(
    meeting_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    meeting = (
        db.query(models.MeetingHistory)
        .filter(
            models.MeetingHistory.id == meeting_id,
            models.MeetingHistory.user_id == current_user.id,
        )
        .first()
    )

    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting MOM not found")

    if not current_user.smtp_setting:
        raise HTTPException(status_code=400, detail="Please connect SMTP first.")

    try:
        workbook = load_workbook(BytesIO(file.file.read()), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")

    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty.")

    headers = [str(h or "").strip().lower() for h in rows[0]]

    if "email" not in headers:
        raise HTTPException(status_code=400, detail="Excel must contain email column.")

    pdf_path = export_to_pdf(
        meeting.generated_mom,
        "meeting",
        f"MOM - {meeting.meeting_title}",
    )

    subject = f"Minutes of Meeting - {meeting.meeting_title}"

    sent = 0
    failed = []

    for row in rows[1:]:
        data = {}

        for index, header in enumerate(headers):
            if index < len(row):
                data[header] = str(row[index] or "").strip()

        email = data.get("email", "")
        if not email:
            continue

        name = data.get("name", "") or "Sir/Madam"
        position = data.get("position", "") or data.get("designation", "")

        try:
            smtp_service.send_email(
                smtp_setting=current_user.smtp_setting,
                to_email=email,
                subject=subject,
                body=build_mom_email(
                    recipient_name=name,
                    position=position,
                    meeting_title=meeting.meeting_title,
                    sender_name=current_user.full_name,
                    sender_designation=current_user.designation,
                ),
                attachments=[pdf_path],
            )
            sent += 1
        except Exception as e:
            failed.append({"email": email, "error": str(e)})

    return {"message": "Excel MOM sending completed.", "sent": sent, "failed": failed}