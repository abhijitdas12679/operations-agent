import json
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.auth import get_current_user
from app.crews.crew_manager import run_report_crew
from app.tools.pdf_tool import export_to_pdf
from app.services.smtp_service import smtp_service

router = APIRouter(prefix="/report", tags=["Report"])


def build_report_email(
    recipient_name: str,
    position: str,
    team_name: str,
    report_date: str,
    sender_name: str | None = None,
    sender_designation: str | None = None,
) -> str:
    name = recipient_name.strip() if recipient_name else "Sir/Madam"
    position_text = f" ({position.strip()})" if position else ""

    signature_name = sender_name or "Team"
    signature_designation = sender_designation or ""

    return f"""Dear {name}{position_text},

Please find attached the daily progress report for {team_name} dated {report_date}.

The attached PDF includes the completed tasks, progress updates, and blockers/challenges for the day.

Kindly review the report.

Best regards,
{signature_name}
{signature_designation}
"""


@router.post("/generate", response_model=schemas.ReportGenerateResponse, status_code=201)
def generate_report(
    req: schemas.ReportGenerateRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    try:
        generated = run_report_crew(
            req.date,
            req.team_name,
            req.tasks_completed,
            req.blockers,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI generation failed: {str(e)}")

    record = models.ReportHistory(
        user_id=current_user.id,
        date=req.date,
        team_name=req.team_name,
        tasks_completed=req.tasks_completed,
        blockers=req.blockers,
        generated_report=generated,
    )

    db.add(record)
    db.commit()
    db.refresh(record)

    return record


@router.get("/history")
def get_report_history(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    return (
        db.query(models.ReportHistory)
        .filter(models.ReportHistory.user_id == current_user.id)
        .order_by(models.ReportHistory.created_at.desc())
        .limit(50)
        .all()
    )


@router.get("/history/{report_id}")
def get_report_detail(
    report_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    record = (
        db.query(models.ReportHistory)
        .filter(
            models.ReportHistory.id == report_id,
            models.ReportHistory.user_id == current_user.id,
        )
        .first()
    )

    if not record:
        raise HTTPException(status_code=404, detail="Report not found")

    return record


@router.post("/send")
def send_report_manual(
    report_id: int = Form(...),
    recipients: str = Form(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    report = (
        db.query(models.ReportHistory)
        .filter(
            models.ReportHistory.id == report_id,
            models.ReportHistory.user_id == current_user.id,
        )
        .first()
    )

    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    if not current_user.smtp_setting:
        raise HTTPException(
            status_code=400,
            detail="Please connect your SMTP email account first.",
        )

    try:
        recipient_list = json.loads(recipients)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid recipients data.")

    valid_recipients = []

    for item in recipient_list:
        name = str(item.get("name", "")).strip()
        position = str(item.get("position", "")).strip()
        email = str(item.get("email", "")).strip()

        if email:
            valid_recipients.append(
                {
                    "name": name or "Sir/Madam",
                    "position": position,
                    "email": email,
                }
            )

    if not valid_recipients:
        raise HTTPException(status_code=400, detail="Please add at least one recipient email.")

    pdf_path = export_to_pdf(
        report.generated_report,
        "report",
        f"Daily Progress Report - {report.team_name} - {report.date}",
    )

    subject = f"Daily Progress Report - {report.team_name} - {report.date}"

    sent = 0
    failed = []

    for recipient in valid_recipients:
        try:
            smtp_service.send_email(
                smtp_setting=current_user.smtp_setting,
                to_email=recipient["email"],
                subject=subject,
                body=build_report_email(
                    recipient_name=recipient["name"],
                    position=recipient["position"],
                    team_name=report.team_name,
                    report_date=report.date,
                    sender_name=current_user.full_name,
                    sender_designation=current_user.designation,
                ),
                attachments=[pdf_path],
            )
            sent += 1
        except Exception as e:
            failed.append(
                {
                    "email": recipient["email"],
                    "name": recipient["name"],
                    "error": str(e),
                }
            )

    return {
        "message": "Report sending completed.",
        "sent": sent,
        "failed": failed,
    }


@router.post("/send-excel")
def send_report_excel(
    report_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    report = (
        db.query(models.ReportHistory)
        .filter(
            models.ReportHistory.id == report_id,
            models.ReportHistory.user_id == current_user.id,
        )
        .first()
    )

    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    if not current_user.smtp_setting:
        raise HTTPException(
            status_code=400,
            detail="Please connect your SMTP email account first.",
        )

    if not file.filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400, detail="Please upload a valid Excel file.")

    try:
        workbook = load_workbook(BytesIO(file.file.read()), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")

    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty.")

    headers = [str(header or "").strip().lower() for header in rows[0]]

    if "email" not in headers:
        raise HTTPException(status_code=400, detail="Excel must contain an email column.")

    pdf_path = export_to_pdf(
        report.generated_report,
        "report",
        f"Daily Progress Report - {report.team_name} - {report.date}",
    )

    subject = f"Daily Progress Report - {report.team_name} - {report.date}"

    sent = 0
    failed = []

    for row in rows[1:]:
        data = {}

        for index, header in enumerate(headers):
            if index < len(row):
                data[header] = str(row[index] or "").strip()

        email = data.get("email", "")
        name = data.get("name", "") or "Sir/Madam"
        position = data.get("position", "") or data.get("designation", "")

        if not email:
            continue

        try:
            smtp_service.send_email(
                smtp_setting=current_user.smtp_setting,
                to_email=email,
                subject=subject,
                body=build_report_email(
                    recipient_name=name,
                    position=position,
                    team_name=report.team_name,
                    report_date=report.date,
                    sender_name=current_user.full_name,
                    sender_designation=current_user.designation,
                ),
                attachments=[pdf_path],
            )
            sent += 1
        except Exception as e:
            failed.append({"email": email, "name": name, "error": str(e)})

    return {
        "message": "Excel report sending completed.",
        "sent": sent,
        "failed": failed,
    }